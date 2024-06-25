from django.contrib import admin, messages
from django.http import HttpResponseRedirect, HttpResponse
from openpyxl.workbook import Workbook

from .forms import TrainerActionForm
from .models import Trainer, Training, Attendance, Price, TrainingSchedule


class PriceInline(admin.TabularInline):
    model = Price
    extra = 1
    fields = ('quantity_to', 'price_to', 'quantity_from', 'price_from')
    verbose_name = "Цена"
    verbose_name_plural = "Цены"


class TrainingScheduleInline(admin.TabularInline):
    model = TrainingSchedule
    extra = 1
    fields = ('day_of_week', 'start_time', 'end_time')
    verbose_name = "Расписание"
    verbose_name_plural = "Расписания"


def export_trainers_data_to_excel(request, trainer_ids=None, start_date=None, end_date=None):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=trainer_data.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = "Тренера"

    columns = ['Имя', 'Фамилия', 'Телефон', 'Занятие', 'Пришли', 'Оплата',
               'Дата', 'Начало', 'Конец']
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = cell.font.copy(bold=True)
        ws.column_dimensions[cell.column_letter].width = 14

    row_num = 1
    trainers = Trainer.objects.all()
    if trainer_ids:
        trainers = trainers.filter(id__in=trainer_ids)

    for trainer in trainers:
        trainings = Training.objects.filter(trainer=trainer)
        for training in trainings:
            attendances = Attendance.objects.filter(training=training)
            if start_date:
                attendances = attendances.filter(recording_date__gte=start_date)
            if end_date:
                attendances = attendances.filter(recording_date__lte=end_date)

            for attendance in attendances:
                schedules = TrainingSchedule.objects.filter(training=training)
                if schedules.exists():
                    schedule = schedules.first()
                    prices = Price.objects.filter(training=training)
                    price = 0
                    for p in prices:
                        if attendance.attend_count <= p.quantity_to:
                            price = p.price_to * attendance.attend_count
                            break
                        elif attendance.attend_count >= p.quantity_from:
                            price = p.price_from * attendance.attend_count

                    row_num += 1
                    row = [
                        trainer.first_name,
                        trainer.last_name,
                        trainer.phone_number,
                        training.name,
                        attendance.attend_count,
                        price,
                        attendance.recording_date.strftime('%Y-%m-%d'),
                        schedule.start_time.strftime('%H:%M'),
                        schedule.end_time.strftime('%H:%M')
                    ]
                    for col_num, cell_value in enumerate(row, 1):
                        ws.cell(row=row_num, column=col_num, value=cell_value)

    wb.save(response)
    return response


@admin.register(Trainer)
class TrainerAdmin(admin.ModelAdmin):
    actions = ['export_trainers']
    action_form = TrainerActionForm

    def export_trainers(self, request, queryset):
        trainer_ids = queryset.values_list('id', flat=True) if queryset.exists() else None
        form = TrainerActionForm(request.POST)
        form.fields['action'].choices = [('export_trainers', self.export_trainers.short_description)]
        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            return export_trainers_data_to_excel(request, trainer_ids=trainer_ids, start_date=start_date,
                                                 end_date=end_date)
        else:
            self.message_user(request, f"Пожалуйста, введите корректные данные: {form.errors.as_text()}",
                              level=messages.ERROR)
            return HttpResponseRedirect(request.get_full_path())

    export_trainers.short_description = 'Экспорт данных тренеров в Excel'

    list_display = ('first_name', 'last_name', 'phone_number')
    search_fields = ('first_name', 'last_name', 'phone_number')


@admin.register(Training)
class TrainingAdmin(admin.ModelAdmin):
    list_display = ('name', 'trainer', 'start_date', 'end_date')
    search_fields = ('name', 'trainer__first_name', 'trainer__last_name')
    list_filter = ('start_date', 'end_date')
    ordering = ('start_date',)
    inlines = [TrainingScheduleInline, PriceInline]


@admin.register(Attendance)
class AttendanceAdmin(admin.ModelAdmin):
    list_display = ('training', 'attend_count', 'recording_day', 'recording_date', 'created_date', 'update_date')
    search_fields = ('training__name', 'recording_day')
    list_filter = ('recording_date', 'recording_day')
    ordering = ('recording_date',)